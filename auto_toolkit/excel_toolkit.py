"""
Excel Automation Toolkit — 一键自动化Excel数据处理
====================================================
功能：批量合并、数据清洗、格式转换、智能分列、报表生成
适用场景：数据分析师、财务人员、运营人员日常Excel处理
"""

import os
import sys
import json
import csv
from datetime import datetime

def _check_file_exists(path: str):
    """Helper: validate file exists, return (True, None) or (False, error_msg)"""
    if not os.path.exists(path):
        return False, f"文件不存在: {path}"
    return True, None

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False


def merge_excel_files(input_dir: str, output_file: str, sheet_name: str = None):
    """合并目录下所有Excel文件为单个文件"""
    if not HAS_PANDAS:
        return {"success": False, "error": "需要安装pandas: pip install pandas"}
    
    if not os.path.isdir(input_dir):
        return {"success": False, "error": f"目录不存在: {input_dir}"}
    
    all_data = []
    files_merged = 0
    
    for f in sorted(os.listdir(input_dir)):
        if f.endswith(('.xlsx', '.xls')) and not f.startswith('~'):
            try:
                filepath = os.path.join(input_dir, f)
                if sheet_name:
                    df = pd.read_excel(filepath, sheet_name=sheet_name)
                else:
                    df = pd.read_excel(filepath)
                all_data.append(df)
                files_merged += 1
            except Exception as e:
                print(f"  跳过 {f}: {e}")
    
    if not all_data:
        return {"success": False, "error": "未找到有效的Excel文件"}
    
    try:
        merged = pd.concat(all_data, ignore_index=True)
        merged.to_excel(output_file, index=False)
    except Exception as e:
        return {"success": False, "error": f"合并失败: {str(e)}"}
    
    return {
        "success": True,
        "files_merged": files_merged,
        "total_rows": len(merged),
        "output": output_file
    }


def clean_excel_data(input_file: str, output_file: str, options: dict = None):
    """清洗Excel数据：去重、去空行、标准化格式"""
    if not HAS_PANDAS:
        return {"success": False, "error": "需要安装pandas: pip install pandas"}
    
    exists, err = _check_file_exists(input_file)
    if not exists:
        return {"success": False, "error": err}
    
    try:
        opts = options or {}
        df = pd.read_excel(input_file)
        original_rows = len(df)
        
        if opts.get('remove_duplicates', True):
            df = df.drop_duplicates()
        
        if opts.get('remove_empty_rows', True):
            df = df.dropna(how='all')
        
        if opts.get('trim_whitespace', True):
            for col in df.select_dtypes(include='object').columns:
                df[col] = df[col].str.strip()
        
        if opts.get('fill_empty', ''):
            df = df.fillna(opts['fill_empty'])
        
        df.to_excel(output_file, index=False)
        
        return {
            "success": True,
            "original_rows": original_rows,
            "cleaned_rows": len(df),
            "removed": original_rows - len(df),
            "output": output_file
        }
    except Exception as e:
        return {"success": False, "error": f"处理失败: {str(e)}"}


def excel_to_csv(input_file: str, output_file: str = None):
    """Excel转CSV"""
    if not HAS_OPENPYXL:
        return {"success": False, "error": "需要安装openpyxl: pip install openpyxl"}
    
    exists, err = _check_file_exists(input_file)
    if not exists:
        return {"success": False, "error": err}
    
    try:
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active
        
        if output_file is None:
            output_file = input_file.rsplit('.', 1)[0] + '.csv'
        
        with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)
        
        return {
            "success": True,
            "output": output_file,
            "rows": ws.max_row,
            "columns": ws.max_column
        }
    except Exception as e:
        return {"success": False, "error": f"转换失败: {str(e)}"}


def csv_to_excel(input_file: str, output_file: str = None):
    """CSV转Excel（带格式化）"""
    if not HAS_OPENPYXL:
        return {"success": False, "error": "需要安装openpyxl: pip install openpyxl"}
    
    exists, err = _check_file_exists(input_file)
    if not exists:
        return {"success": False, "error": err}
    
    try:
        if output_file is None:
            output_file = input_file.rsplit('.', 1)[0] + '.xlsx'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Style for headers
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        with open(input_file, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader, 1):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for col_idx in range(1, ws.max_column + 1):
            max_width = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False):
                for cell in row:
                    if cell.value:
                        max_width = max(max_width, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width + 2, 50)
        
        wb.save(output_file)
        
        return {
            "success": True,
            "output": output_file,
            "rows": ws.max_row,
            "columns": ws.max_column
        }
    except Exception as e:
        return {"success": False, "error": f"转换失败: {str(e)}"}


def split_by_column(input_file: str, column: str, output_dir: str = None):
    """按指定列的值拆分Excel为多个文件"""
    if not HAS_PANDAS:
        return {"success": False, "error": "需要安装pandas: pip install pandas"}
    
    exists, err = _check_file_exists(input_file)
    if not exists:
        return {"success": False, "error": err}
    
    try:
        df = pd.read_excel(input_file)
        if column not in df.columns:
            return {"success": False, "error": f"列 '{column}' 不存在。可用列: {list(df.columns)}"}
        
        if output_dir is None:
            output_dir = os.path.splitext(input_file)[0] + '_split'
        
        os.makedirs(output_dir, exist_ok=True)
        groups = df.groupby(column)
        results = []
        
        for name, group in groups:
            safe_name = str(name).replace('/', '_').replace('\\', '_')[:50]
            output_file = os.path.join(output_dir, f'{safe_name}.xlsx')
            group.to_excel(output_file, index=False)
            results.append({"group": safe_name, "rows": len(group), "file": output_file})
        
        return {
            "success": True,
            "total_groups": len(results),
            "output_dir": output_dir,
            "groups": results
        }
    except Exception as e:
        return {"success": False, "error": f"拆分失败: {str(e)}"}


def generate_summary_report(input_file: str, output_file: str = None):
    """生成数据摘要报表"""
    if not HAS_PANDAS:
        return {"success": False, "error": "需要安装pandas: pip install pandas"}
    
    exists, err = _check_file_exists(input_file)
    if not exists:
        return {"success": False, "error": err}
    
    try:
        df = pd.read_excel(input_file)
    
        report = {
            "file": input_file,
            "generated_at": datetime.now().isoformat(),
            "shape": {"rows": len(df), "columns": len(df.columns)},
            "columns": {},
            "missing_data": {}
        }
    
        for col in df.columns:
            col_data = df[col]
            col_info = {"dtype": str(col_data.dtype)}
            
            missing = col_data.isna().sum()
            report["missing_data"][col] = f"{missing} ({missing/len(df)*100:.1f}%)"
            
            if pd.api.types.is_numeric_dtype(col_data):
                col_info.update({
                    "min": float(col_data.min()) if not col_data.isna().all() else None,
                    "max": float(col_data.max()) if not col_data.isna().all() else None,
                    "mean": round(float(col_data.mean()), 2) if not col_data.isna().all() else None,
                    "sum": float(col_data.sum()) if not col_data.isna().all() else None,
                })
            else:
                unique = col_data.nunique()
                col_info["unique_values"] = int(unique)
                if unique <= 10:
                    top = col_data.value_counts().head(5).to_dict()
                    col_info["top_values"] = {str(k): int(v) for k, v in top.items()}
            
            report["columns"][col] = col_info
        
        # Save report
        report_file = output_file or input_file.rsplit('.', 1)[0] + '_report.json'
        with open(report_file, 'w', encoding='utf-8') as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        
        # Also generate formatted Excel report
        if output_file is None:
            xlsx_report = input_file.rsplit('.', 1)[0] + '_report.xlsx'
        else:
            xlsx_report = output_file
        
        if HAS_OPENPYXL:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "数据摘要"
            
            # Write report
            ws['A1'] = "数据摘要报表"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f"文件: {os.path.basename(input_file)}"
            ws['A3'] = f"行数: {len(df)}, 列数: {len(df.columns)}"
            ws['A4'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            row = 6
            ws.cell(row=row, column=1, value="列名").font = Font(bold=True)
            ws.cell(row=row, column=2, value="类型").font = Font(bold=True)
            ws.cell(row=row, column=3, value="缺失值").font = Font(bold=True)
            ws.cell(row=row, column=4, value="统计信息").font = Font(bold=True)
            
            for col, info in report["columns"].items():
                row += 1
                ws.cell(row=row, column=1, value=col)
                ws.cell(row=row, column=2, value=info.get("dtype", ""))
                ws.cell(row=row, column=3, value=report["missing_data"].get(col, ""))
                
                stats = []
                if "min" in info: stats.append(f"min={info['min']}")
                if "max" in info: stats.append(f"max={info['max']}")
                if "mean" in info: stats.append(f"mean={info['mean']}")
                if "unique_values" in info: stats.append(f"unique={info['unique_values']}")
                ws.cell(row=row, column=4, value=", ".join(stats))
            
            wb.save(xlsx_report)
        
        report["report_file"] = report_file
        report["xlsx_report"] = xlsx_report
        
        return {"success": True, "data": report}
    except Exception as e:
        return {"success": False, "error": f"生成报表失败: {str(e)}"}


def interactive_menu():
    """命令行交互菜单"""
    print("=" * 60)
    print("   Excel Automation Toolkit — 自动化Excel处理工具箱")
    print("=" * 60)
    print()
    print("  1. 合并多个Excel文件")
    print("  2. 清洗Excel数据（去重/去空/标准化）")
    print("  3. Excel → CSV 转换")
    print("  4. CSV → Excel（带格式化）")
    print("  5. 按列拆分Excel为多个文件")
    print("  6. 生成数据摘要报表")
    print("  0. 退出")
    print()
    
    choice = input("请选择功能 [0-6]: ").strip()
    
    if choice == '0':
        print("再见！")
        return
    
    functions = {
        '1': ('合并Excel文件', merge_excel_files_ui),
        '2': ('清洗Excel数据', clean_excel_ui),
        '3': ('Excel转CSV', excel_to_csv_ui),
        '4': ('CSV转Excel', csv_to_excel_ui),
        '5': ('按列拆分Excel', split_excel_ui),
        '6': ('生成数据摘要报表', report_ui),
    }
    
    if choice in functions:
        name, fn = functions[choice]
        print(f"\n--- {name} ---")
        try:
            result = fn()
            if result.get('success'):
                print(f"\n✓ 操作完成！")
                for k, v in result.items():
                    if k != 'success' and k != 'data':
                        print(f"  {k}: {v}")
            else:
                print(f"\n✗ 操作失败: {result.get('error', '未知错误')}")
        except Exception as e:
            print(f"\n✗ 出错: {e}")
    else:
        print("无效选择")
    
    input("\n按回车继续...")
    interactive_menu()


def merge_excel_files_ui():
    d = input("Excel文件目录: ").strip()
    o = input("输出文件 [merged.xlsx]: ").strip() or "merged.xlsx"
    return merge_excel_files(d, o)

def clean_excel_ui():
    f = input("输入Excel文件: ").strip()
    o = input("输出文件 [cleaned.xlsx]: ").strip() or "cleaned.xlsx"
    return clean_excel_data(f, o)

def excel_to_csv_ui():
    f = input("输入Excel文件: ").strip()
    o = input("输出CSV文件 [留空自动命名]: ").strip() or None
    return excel_to_csv(f, o)

def csv_to_excel_ui():
    f = input("输入CSV文件: ").strip()
    o = input("输出Excel文件 [留空自动命名]: ").strip() or None
    return csv_to_excel(f, o)

def split_excel_ui():
    f = input("输入Excel文件: ").strip()
    c = input("按哪一列拆分: ").strip()
    o = input("输出目录 [留空自动创建]: ").strip() or None
    return split_by_column(f, c, o)

def report_ui():
    f = input("输入Excel文件: ").strip()
    o = input("输出Excel报表 [留空自动命名]: ").strip() or None
    return generate_summary_report(f, o)


if __name__ == '__main__':
    if len(sys.argv) > 1:
        # Command-line mode
        cmd = sys.argv[1]
        if cmd == 'merge' and len(sys.argv) >= 4:
            print(json.dumps(merge_excel_files(sys.argv[2], sys.argv[3]), ensure_ascii=False))
        elif cmd == 'clean' and len(sys.argv) >= 4:
            print(json.dumps(clean_excel_data(sys.argv[2], sys.argv[3]), ensure_ascii=False))
        elif cmd == 'xls2csv' and len(sys.argv) >= 3:
            out = sys.argv[3] if len(sys.argv) > 3 else None
            print(json.dumps(excel_to_csv(sys.argv[2], out), ensure_ascii=False))
        elif cmd == 'csv2xls' and len(sys.argv) >= 3:
            out = sys.argv[3] if len(sys.argv) > 3 else None
            print(json.dumps(csv_to_excel(sys.argv[2], out), ensure_ascii=False))
        elif cmd == 'split' and len(sys.argv) >= 4:
            out_dir = sys.argv[4] if len(sys.argv) > 4 else None
            print(json.dumps(split_by_column(sys.argv[2], sys.argv[3], out_dir), ensure_ascii=False))
        elif cmd == 'report' and len(sys.argv) >= 3:
            out = sys.argv[3] if len(sys.argv) > 3 else None
            print(json.dumps(generate_summary_report(sys.argv[2], out), ensure_ascii=False))
        else:
            print("用法: python excel_toolkit.py <命令> [参数...]")
            print("  merge <目录> <输出>       合并Excel")
            print("  clean <输入> <输出>       清洗数据")
            print("  xls2csv <输入> [输出]     Excel转CSV")
            print("  csv2xls <输入> [输出]     CSV转Excel")
            print("  split <输入> <列> [目录]  按列拆分")
            print("  report <输入> [输出]      生成报表")
    else:
        interactive_menu()
